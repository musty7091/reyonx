from flask import Blueprint, render_template, request, redirect, current_app, send_file
from database import db
from models import Sale, Waste, Expense, Period, SystemSetting, SaleItem, Product, Invoice, Payment
from datetime import datetime
from decimal import Decimal
from sqlalchemy import func
import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

report_bp = Blueprint("report", __name__)

def get_active_period():
    period = Period.query.filter_by(is_active=True).first()
    if not period:
        period = Period(name="1. Dönem")
        db.session.add(period)
        db.session.commit()
    return period

def get_bonus_rate():
    rate_setting = SystemSetting.query.filter_by(setting_key="bonus_rate").first()
    return Decimal(rate_setting.setting_value) if rate_setting else Decimal("5")

def calculate_debt_at_time(end_time):
    # İptal edilen kayıtları SQL sorgusunda harici tutuyoruz (is_cancelled == False)
    query_inv_alis = db.session.query(func.sum(Invoice.total_amount)).filter(Invoice.invoice_type == "alis", Invoice.is_cancelled == False)
    query_inv_iade = db.session.query(func.sum(Invoice.total_amount)).filter(Invoice.invoice_type == "iade", Invoice.is_cancelled == False)
    query_pay = db.session.query(func.sum(Payment.amount)).filter(Payment.is_cancelled == False)
    
    if end_time:
        query_inv_alis = query_inv_alis.filter(Invoice.date <= end_time)
        query_inv_iade = query_inv_iade.filter(Invoice.date <= end_time)
        query_pay = query_pay.filter(Payment.date <= end_time)
        
    t_alis = query_inv_alis.scalar() or Decimal('0.00')
    t_iade = query_inv_iade.scalar() or Decimal('0.00')
    t_paid = query_pay.scalar() or Decimal('0.00')
    
    return t_alis - t_iade - t_paid

@report_bp.route("/report")
def report():
    active_period = get_active_period()
    
    # Rapor ekranında veritabanı toplamlarını çekerken iptal edilenleri saymıyoruz
    total_revenue = db.session.query(func.sum(Sale.total_revenue)).filter(Sale.period_id == active_period.id, Sale.is_cancelled == False).scalar() or Decimal('0.00')
    total_cost = db.session.query(func.sum(Sale.total_cost)).filter(Sale.period_id == active_period.id, Sale.is_cancelled == False).scalar() or Decimal('0.00')
    total_waste_cost = db.session.query(func.sum(Waste.quantity * Waste.cost)).filter(Waste.period_id == active_period.id, Waste.is_cancelled == False).scalar() or Decimal('0.00')
    total_expenses = db.session.query(func.sum(Expense.amount)).filter(Expense.period_id == active_period.id, Expense.is_cancelled == False).scalar() or Decimal('0.00')

    gross_profit = total_revenue - total_cost
    net_profit = gross_profit - total_waste_cost - total_expenses
    
    bonus_rate = get_bonus_rate()
    bonus = (net_profit * (bonus_rate / Decimal("100"))) if net_profit > Decimal('0') else Decimal('0.00')
    
    current_debt = calculate_debt_at_time(None)

    wastes = Waste.query.filter_by(period_id=active_period.id, is_cancelled=False).all()

    return render_template(
        "profit_report.html",
        active_period=active_period,
        total_revenue=total_revenue,
        total_cost=total_cost,
        gross_profit=gross_profit,
        total_waste_cost=total_waste_cost,
        total_expenses=total_expenses,
        net_profit=net_profit,
        bonus=bonus,
        bonus_rate=bonus_rate,
        wastes=wastes,
        current_debt=current_debt
    )

@report_bp.route("/period/close", methods=["POST"])
def close_period():
    active_period = get_active_period()
    
    t_rev = db.session.query(func.sum(Sale.total_revenue)).filter(Sale.period_id == active_period.id, Sale.is_cancelled == False).scalar() or Decimal('0.00')
    t_cost = db.session.query(func.sum(Sale.total_cost)).filter(Sale.period_id == active_period.id, Sale.is_cancelled == False).scalar() or Decimal('0.00')
    t_waste = db.session.query(func.sum(Waste.quantity * Waste.cost)).filter(Waste.period_id == active_period.id, Waste.is_cancelled == False).scalar() or Decimal('0.00')
    t_exp = db.session.query(func.sum(Expense.amount)).filter(Expense.period_id == active_period.id, Expense.is_cancelled == False).scalar() or Decimal('0.00')
    
    n_prof = (t_rev - t_cost) - t_waste - t_exp

    active_period.total_revenue = t_rev
    active_period.total_cost = t_cost
    active_period.total_waste_cost = t_waste
    active_period.total_expenses = t_exp
    active_period.net_profit = n_prof
    
    active_period.bonus_rate = get_bonus_rate() 
    
    active_period.end_date = datetime.utcnow()
    active_period.is_active = False 

    new_name = request.form.get("new_period_name", "Yeni Dönem")
    new_period = Period(name=new_name)
    db.session.add(new_period)
    db.session.commit()

    return redirect("/periods")

@report_bp.route("/periods")
def periods():
    page = request.args.get('page', 1, type=int)
    periods_pagination = Period.query.order_by(Period.id.desc()).paginate(page=page, per_page=10, error_out=False)
    
    return render_template("periods.html", pagination=periods_pagination)

@report_bp.route("/period/archive/<int:id>")
def view_archive(id):
    period = Period.query.get_or_404(id)
    
    bonus_rate = period.bonus_rate if period.bonus_rate is not None else get_bonus_rate()
    bonus = (period.net_profit * (bonus_rate / Decimal("100"))) if period.net_profit > Decimal('0') else Decimal('0.00')
    
    period_debt = calculate_debt_at_time(period.end_date)
    
    sold_items = db.session.query(
        Product.name,
        func.sum(SaleItem.quantity).label('total_quantity')
    ).join(SaleItem, Product.id == SaleItem.product_id)\
     .join(Sale, Sale.id == SaleItem.sale_id)\
     .filter(Sale.period_id == period.id, Sale.is_cancelled == False)\
     .group_by(Product.id).order_by(func.sum(SaleItem.quantity).desc()).all()
    
    return render_template(
        "closed_report.html", 
        period=period, 
        bonus=bonus,
        bonus_rate=bonus_rate,
        sold_items=sold_items,
        period_debt=period_debt
    )

@report_bp.route("/period/export/<int:id>")
def export_excel(id):
    period = Period.query.get_or_404(id)
    
    bonus_rate = period.bonus_rate if period.bonus_rate is not None else get_bonus_rate()
    bonus = (period.net_profit * (bonus_rate / Decimal("100"))) if period.net_profit > Decimal('0') else Decimal('0.00')
    
    period_debt = calculate_debt_at_time(period.end_date)

    sold_items = db.session.query(
        Product.name,
        func.sum(SaleItem.quantity).label('total_quantity')
    ).join(SaleItem, Product.id == SaleItem.product_id)\
     .join(Sale, Sale.id == SaleItem.sale_id)\
     .filter(Sale.period_id == period.id, Sale.is_cancelled == False)\
     .group_by(Product.id).order_by(func.sum(SaleItem.quantity).desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Dönem Raporu"

    bold_font = Font(bold=True)
    center_aligned_text = Alignment(horizontal="center")

    ws.merge_cells('A1:B1')
    ws['A1'] = "REYONX İŞLETME RAPORU"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center_aligned_text

    ws.append(["Dönem Adı:", period.name])
    ws.append(["Kapanış Tarihi:", period.end_date.strftime('%d.%m.%Y') if period.end_date else '-'])
    ws.append([])

    ws.append(["FİNANSAL ÖZET", ""])
    ws[f'A{ws.max_row}'].font = bold_font
    
    ws.append(["Toplam Satış (Ciro):", f"{period.total_revenue:.2f} ₺"])
    ws.append(["Satılan Malın Maliyeti:", f"-{period.total_cost:.2f} ₺"])
    ws.append(["Brüt Kâr:", f"{(period.total_revenue - period.total_cost):.2f} ₺"])
    ws.append(["Fire ve Kayıp Giderleri:", f"-{period.total_waste_cost:.2f} ₺"])
    ws.append(["İşletme Giderleri:", f"-{period.total_expenses:.2f} ₺"])
    
    ws.append(["GERÇEK NET KÂR:", f"{period.net_profit:.2f} ₺"])
    ws[f'A{ws.max_row}'].font = bold_font
    ws[f'B{ws.max_row}'].font = bold_font
    
    ws.append([f"Hesaplanan Personel Primi (%{bonus_rate}):", f"{bonus:.2f} ₺"])
    
    ws.append(["DÖNEM SONU TİCARİ BORÇ:", f"{period_debt:.2f} ₺"])
    ws[f'A{ws.max_row}'].font = bold_font
    ws[f'B{ws.max_row}'].font = bold_font
    
    ws.append([]) 

    ws.append(["BU DÖNEM SATILAN ÜRÜNLER", ""])
    ws[f'A{ws.max_row}'].font = bold_font
    
    ws.append(["Ürün Adı", "Satılan Adet"])
    ws[f'A{ws.max_row}'].font = bold_font
    ws[f'B{ws.max_row}'].font = bold_font

    if not sold_items:
        ws.append(["Satış bulunamadı.", "0"])
    else:
        for item in sold_items:
            ws.append([item.name, item.total_quantity])

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    
    return send_file(
        out, 
        download_name=f"ReyonX_Rapor_{period.name}.xlsx", 
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )